#!/usr/bin/env python3
"""
从 WPS 源文件重建 Excel 2021 兼容的索具库存表
- 仅复制数据，不复制可能不兼容的格式
- 添加物料-供应商级联下拉（前50行）
"""
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def copy_sheet_data(src_ws, dst_ws):
    """复制工作表所有数据（含公式），保持完整结构"""
    for r in range(1, src_ws.max_row + 1):
        for c in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=r, column=c)
            val = src_cell.value
            # ArrayFormula 等取公式字符串
            if hasattr(val, 'text'):
                val = val.text
            dst_cell = dst_ws.cell(row=r, column=c)
            dst_cell.value = val if val is not None else ""  # 空值用""保留行结构

def main():
    from openpyxl import load_workbook
    
    print("读取源文件...")
    src = load_workbook('索具库存表(1).xlsx', data_only=False)
    
    print("创建新工作簿（Excel 2021 兼容）...")
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 sheet
    
    # 1. 复制各 sheet（Sheet1 仅创建空表，不复制 WPS 可能产生的冗余数据）
    for name in ['实时库存', '入库明细', '出库明细', '物料表']:
        src_ws = src[name]
        ws = wb.create_sheet(name)
        copy_sheet_data(src_ws, ws)
        print(f"  已复制: {name} (源{src_ws.max_row}行 -> 目标{ws.max_row}行)")
    ws = wb.create_sheet('Sheet1')  # 新建空 Sheet1，用于级联辅助
    
    src.close()
    
    cm = wb['出库明细']
    wl = wb['物料表']
    sh1 = wb['Sheet1']
    
    wl_start, wl_end = 3, min(474, wl.max_row)
    cm_start, cm_cascade_end = 3, 52
    cm_full_end = min(509, cm.max_row)
    
    # 2. Sheet1：级联辅助公式（前50行，全新创建无 WPS 残留）
    for r in range(cm_start, cm_cascade_end + 1):
        for col in range(1, 11):
            formula = (
                f'=IFERROR(INDEX(物料表!$F$3:$F${wl_end},'
                f'AGGREGATE(15,6,(ROW(物料表!$B$3:$B${wl_end})-2)/(物料表!$B$3:$B${wl_end}=出库明细!$B${r}),{col})),"")'
            )
            sh1.cell(row=r, column=col, value=formula)
    
    # 4. 出库明细：移除可能存在的公式，准备重新设置
    for r in range(cm_start, cm_full_end + 1):
        for col in [3, 4, 5, 6]:  # C,D,E,F
            cell = cm.cell(row=r, column=col)
            if isinstance(cell.value, str) and str(cell.value).startswith('='):
                cell.value = None
    
    # 5. 出库明细 B 列：物料下拉
    cm.data_validations.dataValidation = []  # 清空所有验证，重新添加
    dv_b = DataValidation(
        type="list",
        formula1=f'物料表!$B$3:$B${wl_end}',
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True
    )
    dv_b.add(f'B1:B{cm_full_end}')
    cm.add_data_validation(dv_b)
    
    # 6. 出库明细 F 列：供应商级联下拉（前50行）
    for r in range(cm_start, cm_cascade_end + 1):
        dv = DataValidation(
            type="list",
            formula1=f'Sheet1!A{r}:J{r}',
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        dv.add(f'F{r}')
        cm.add_data_validation(dv)
    
    # 7. 出库明细 F 列：第51行起用公式取第一个供应商
    for r in range(cm_cascade_end + 1, cm_full_end + 1):
        formula = f'=IFERROR(INDEX(物料表!$F$3:$F${wl_end},MATCH(B{r},物料表!$B$3:$B${wl_end},0)),"")'
        cm.cell(row=r, column=6, value=formula)
    
    # 8. 出库明细 C/D/E 列：公式
    for col_idx, wl_col in [(3, 'C'), (4, 'D'), (5, 'E')]:
        for r in range(cm_start, cm_cascade_end + 1):
            formula = (
                f'=IFERROR(INDEX(物料表!${wl_col}${wl_start}:${wl_col}${wl_end},'
                f'AGGREGATE(15,6,(ROW(物料表!$B$3:$B${wl_end})-2)/((物料表!$B$3:$B${wl_end}=B{r})*(物料表!$F$3:$F${wl_end}=F{r})),1)),"")'
            )
            cm.cell(row=r, column=col_idx, value=formula)
        for r in range(cm_cascade_end + 1, cm_full_end + 1):
            formula = f'=IFERROR(INDEX(物料表!${wl_col}${wl_start}:${wl_col}${wl_end},MATCH(B{r},物料表!$B$3:$B${wl_end},0)),"")'
            cm.cell(row=r, column=col_idx, value=formula)
    
    out_file = '索具库存表_Excel2021版.xlsx'
    wb.save(out_file)
    print(f"\n已生成: {out_file}")
    print("说明: 全新创建，兼容 Excel 2021，第3-52行支持物料-供应商级联下拉")

if __name__ == '__main__':
    main()
