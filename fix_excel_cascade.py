#!/usr/bin/env python3
"""
修改索具库存表：实现出库明细中物料与供应商的级联选择
- 精简方案：仅支持前100行，使用固定范围验证，避免 INDIRECT 和大量公式
- 兼容 Excel 2021，避免触发「修复不可读取内容」
"""
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def main():
    wb = load_workbook('索具库存表(1).xlsx', data_only=False)
    
    cm = wb['出库明细']
    wl = wb['物料表']
    sh1 = wb['Sheet1']
    
    wl_start, wl_end = 3, 474
    cm_start = 3
    cm_cascade_end = 52    # 级联下拉支持到第52行（共50行），避免过多验证导致修复提示
    cm_full_end = 509
    max_suppliers = 10
    
    # 0. 清空 Sheet1 中我们之前可能添加的所有内容（避免残留导致问题）
    for row in sh1.iter_rows(min_row=1, max_row=600, min_col=1, max_col=15):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = None
    
    # 1. Sheet1：仅前100行有公式（每行 A-J 列）
    for r in range(cm_start, cm_cascade_end + 1):
        for col in range(1, max_suppliers + 1):
            formula = (
                f'=IFERROR(INDEX(物料表!$F$3:$F${wl_end},'
                f'AGGREGATE(15,6,(ROW(物料表!$B$3:$B${wl_end})-2)/(物料表!$B$3:$B${wl_end}=出库明细!$B${r}),{col})),"")'
            )
            sh1.cell(row=r, column=col, value=formula)
    
    # 2. 清空 F 列公式
    for r in range(cm_start, cm_full_end + 1):
        cell = cm.cell(row=r, column=6)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            cell.value = None
    
    # 3. 移除所有旧的 F 列验证
    cm.data_validations.dataValidation = [
        dv for dv in cm.data_validations.dataValidation
        if 'F' not in str(dv.sqref)
    ]
    
    # 4. F 列：仅前100行用固定范围验证（无 INDIRECT，每行一个验证对象）
    for r in range(cm_start, cm_cascade_end + 1):
        dv = DataValidation(
            type="list",
            formula1=f'Sheet1!A{r}:J{r}',  # 固定范围，无公式
            allow_blank=True,
            showInputMessage=True,
            showErrorMessage=True
        )
        dv.add(f'F{r}')
        cm.add_data_validation(dv)
    
    # 5. 第101-509行：F 列用简单公式取第一个供应商（无下拉）
    for r in range(cm_cascade_end + 1, cm_full_end + 1):
        formula = f'=IFERROR(INDEX(物料表!$F$3:$F${wl_end},MATCH(B{r},物料表!$B$3:$B${wl_end},0)),"")'
        cm.cell(row=r, column=6, value=formula)
    
    # 6. C/D/E 列公式
    for col_idx, wl_col in [(3, 'C'), (4, 'D'), (5, 'E')]:
        for r in range(cm_start, cm_cascade_end + 1):
            formula = (
                f'=IFERROR(INDEX(物料表!${wl_col}${wl_start}:${wl_col}${wl_end},'
                f'AGGREGATE(15,6,(ROW(物料表!$B$3:$B${wl_end})-2)/((物料表!$B$3:$B${wl_end}=B{r})*(物料表!$F$3:$F${wl_end}=F{r})),1)),"")'
            )
            cm.cell(row=r, column=col_idx, value=formula)
        for r in range(cm_cascade_end + 1, cm_full_end + 1):
            # 101行起：仅按 B 列匹配（取第一个）
            formula = f'=IFERROR(INDEX(物料表!${wl_col}${wl_start}:${wl_col}${wl_end},MATCH(B{r},物料表!$B$3:$B${wl_end},0)),"")'
            cm.cell(row=r, column=col_idx, value=formula)
    
    out_file = '索具库存表(1)_级联版.xlsx'
    wb.save(out_file)
    print(f'已保存到: {out_file}')
    print('')
    print('修改说明（精简方案，兼容 Excel 2021）:')
    print('  - 第3-52行：B列选物料后，F列可下拉选该物料的所有供应商')
    print('  - 第53-509行：F列自动带出第一个供应商（可手动修改）')
    print('  - 已避免 INDIRECT 和过多公式，减少「修复不可读取内容」风险')

if __name__ == '__main__':
    main()
